import json
import sqlite3
import os
import csv
import uuid
from datetime import datetime, date, timedelta
from io import BytesIO
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'sst_gestor_secreto_2026'
CORS(app)

# ------------------- CONFIGURACIÓN DE ARCHIVOS -------------------
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'jpg', 'jpeg', 'png', 'txt', 'zip', 'rar'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generar_codigo_acceso():
    """Genera un código único de acceso para el cliente"""
    return str(uuid.uuid4())[:8].upper()

# ------------------- BASE DE DATOS -------------------
def init_db():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS actividades (
        codigo TEXT PRIMARY KEY,
        descripcion TEXT,
        nivel_riesgo TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        razon_social TEXT,
        ruc TEXT UNIQUE,
        representante TEXT,
        email TEXT,
        telefono TEXT,
        sector TEXT,
        actividad_codigo TEXT,
        numero_trabajadores INTEGER,
        tiene_grupos_prioritarios INTEGER DEFAULT 0,
        tiene_responsable_previo INTEGER DEFAULT 0,
        fecha_registro TEXT,
        codigo_acceso TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS propuestas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        fecha_creacion TEXT,
        fecha_actualizacion TEXT,
        estado TEXT,
        total_mensual REAL,
        total_anual REAL,
        observaciones TEXT,
        observaciones_personalizadas TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS servicios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        propuesta_id INTEGER,
        descripcion TEXT,
        tipo TEXT,
        cantidad INTEGER,
        precio_unitario REAL,
        subtotal REAL,
        observaciones TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS proyectos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        propuesta_id INTEGER,
        nombre TEXT,
        fecha_inicio TEXT,
        fecha_fin TEXT,
        estado TEXT,
        consultor_responsable TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS entregables (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        proyecto_id INTEGER,
        servicio_id INTEGER,
        nombre TEXT,
        descripcion TEXT,
        estado TEXT,
        fecha_limite TEXT,
        fecha_entrega TEXT,
        archivo TEXT,
        comentarios TEXT,
        es_autogenerado INTEGER DEFAULT 1
    )''')
    
    # Migraciones
    c.execute("PRAGMA table_info(clientes)")
    columnas_clientes = [col[1] for col in c.fetchall()]
    if 'codigo_acceso' not in columnas_clientes:
        c.execute("ALTER TABLE clientes ADD COLUMN codigo_acceso TEXT")
        # Generar códigos para clientes existentes
        c.execute("SELECT id FROM clientes")
        clientes = c.fetchall()
        for cliente in clientes:
            codigo = generar_codigo_acceso()
            c.execute("UPDATE clientes SET codigo_acceso=? WHERE id=?", (codigo, cliente[0]))
        print("✅ Columna 'codigo_acceso' agregada y códigos generados")
    
    c.execute("PRAGMA table_info(propuestas)")
    columnas_prop = [col[1] for col in c.fetchall()]
    if 'fecha_actualizacion' not in columnas_prop:
        c.execute("ALTER TABLE propuestas ADD COLUMN fecha_actualizacion TEXT")
    if 'observaciones_personalizadas' not in columnas_prop:
        c.execute("ALTER TABLE propuestas ADD COLUMN observaciones_personalizadas TEXT")
    
    c.execute("PRAGMA table_info(servicios)")
    columnas_serv = [col[1] for col in c.fetchall()]
    if 'observaciones' not in columnas_serv:
        c.execute("ALTER TABLE servicios ADD COLUMN observaciones TEXT")
    
    c.execute("PRAGMA table_info(proyectos)")
    columnas_proy = [col[1] for col in c.fetchall()]
    if 'propuesta_id' not in columnas_proy:
        c.execute("ALTER TABLE proyectos ADD COLUMN propuesta_id INTEGER")
    
    c.execute("PRAGMA table_info(entregables)")
    columnas_ent = [col[1] for col in c.fetchall()]
    if 'servicio_id' not in columnas_ent:
        c.execute("ALTER TABLE entregables ADD COLUMN servicio_id INTEGER")
    if 'es_autogenerado' not in columnas_ent:
        c.execute("ALTER TABLE entregables ADD COLUMN es_autogenerado INTEGER DEFAULT 1")
    if 'descripcion' not in columnas_ent:
        c.execute("ALTER TABLE entregables ADD COLUMN descripcion TEXT")
    if 'archivo' not in columnas_ent:
        c.execute("ALTER TABLE entregables ADD COLUMN archivo TEXT")
    
    conn.commit()
    conn.close()
    print("✅ Base de datos inicializada correctamente")

def cargar_anexo2_csv():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM actividades")
    if c.fetchone()[0] == 0:
        csv_path = 'data/anexo2.csv'
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    c.execute("INSERT INTO actividades (codigo, descripcion, nivel_riesgo) VALUES (?,?,?)",
                              (row['codigo'], row['descripcion'], row['nivel_riesgo']))
            conn.commit()
            print("✅ Datos del Anexo 2 cargados desde CSV")
        else:
            actividades = [
                ('A0111.11', 'Cultivo de trigo.', 'Bajo'),
                ('B0510.00', 'Extracción de carbón de piedra', 'Alto'),
                ('F4100.10', 'Construcción de edificios', 'Alto'),
                ('G4510.01', 'Venta de vehículos', 'Bajo'),
                ('J5811.01', 'Publicación de libros', 'Medio'),
            ]
            c.executemany("INSERT INTO actividades (codigo, descripcion, nivel_riesgo) VALUES (?,?,?)", actividades)
            conn.commit()
            print("✅ Datos de ejemplo cargados (no se encontró anexo2.csv)")
    conn.close()

def recalcular_totales_anuales():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id FROM propuestas")
    propuestas = c.fetchall()
    for prop_id in propuestas:
        pid = prop_id[0]
        c.execute("SELECT SUM(subtotal) FROM servicios WHERE propuesta_id=? AND tipo='Setup'", (pid,))
        total_setup = c.fetchone()[0] or 0
        c.execute("SELECT SUM(subtotal) FROM servicios WHERE propuesta_id=? AND tipo='Mensual'", (pid,))
        total_mensual = c.fetchone()[0] or 0
        c.execute("SELECT SUM(subtotal) FROM servicios WHERE propuesta_id=? AND tipo='Especial'", (pid,))
        total_especial = c.fetchone()[0] or 0
        total_anual = total_setup + total_mensual + total_especial
        c.execute("UPDATE propuestas SET total_anual=? WHERE id=?", (total_anual, pid))
    conn.commit()
    conn.close()
    print("✅ Totales anuales recalculados como suma simple")

init_db()
cargar_anexo2_csv()
recalcular_totales_anuales()

# ------------------- FUNCIONES DE NEGOCIO -------------------
def obtener_nivel_riesgo(codigo):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT nivel_riesgo FROM actividades WHERE codigo=?", (codigo,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 'Bajo'

def calcular_responsable_y_horas(num_trabajadores, nivel_riesgo):
    if num_trabajadores <= 9:
        if nivel_riesgo in ['Bajo', 'Medio']:
            responsable = 'Monitor'
            horas = 8
        else:
            responsable = 'Técnico'
            horas = 16
        requiere_reglamento = False
    elif 10 <= num_trabajadores <= 49:
        if nivel_riesgo == 'Bajo':
            responsable = 'Monitor'
            horas = 16
        elif nivel_riesgo == 'Medio':
            responsable = 'Técnico'
            horas = 32
        else:
            responsable = 'Técnico'
            horas = 48
        requiere_reglamento = True
    elif 50 <= num_trabajadores <= 99:
        responsable = 'Técnico'
        horas = 32 if nivel_riesgo == 'Bajo' else 48 if nivel_riesgo == 'Medio' else 64
        requiere_reglamento = True
    elif 100 <= num_trabajadores <= 199:
        responsable = 'Técnico'
        horas = 160
        requiere_reglamento = True
    else:
        responsable = 'Técnico'
        horas = 160
        requiere_reglamento = True

    if num_trabajadores >= 50:
        organismo = 'Comité'
    elif 10 <= num_trabajadores <= 49:
        organismo = 'Delegado'
    else:
        organismo = 'Ninguno'

    return responsable, horas, requiere_reglamento, organismo

def generar_servicios_base(cliente):
    nivel = obtener_nivel_riesgo(cliente['actividad_codigo'])
    responsable, horas, requiere_reglamento, organismo = calcular_responsable_y_horas(
        cliente['numero_trabajadores'], nivel
    )
    num_trab = cliente['numero_trabajadores']
    es_micro = num_trab <= 9
    es_pequena_o_mas = num_trab >= 10

    servicios = []

    # ---- SETUP ----
    if es_pequena_o_mas:
        servicios.append({
            'descripcion': 'Elaboración y aprobación del Reglamento de Higiene y Seguridad',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 350.0,
            'observaciones': ''
        })
    else:
        servicios.append({
            'descripcion': 'Elaboración y aprobación del Plan de Prevención de Riesgos Laborales',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 250.0,
            'observaciones': ''
        })

    servicios.append({
        'descripcion': 'Formulación de la Política de Seguridad y Salud en el Trabajo',
        'tipo': 'Setup',
        'cantidad': 1,
        'precio_unitario': 120.0,
        'observaciones': ''
    })

    servicios.append({
        'descripcion': f'Designación de {responsable} de SST y registro en plataforma',
        'tipo': 'Setup',
        'cantidad': 1,
        'precio_unitario': 100.0,
        'observaciones': ''
    })

    if organismo != 'Ninguno':
        servicios.append({
            'descripcion': f'Conformación y registro de {organismo} de SST',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 180.0,
            'observaciones': ''
        })

    servicios.append({
        'descripcion': 'Implementación y registro de Sala de Apoyo a la Lactancia Materna (si aplica)',
        'tipo': 'Setup',
        'cantidad': 1,
        'precio_unitario': 80.0,
        'observaciones': ''
    })

    if es_micro:
        servicios.append({
            'descripcion': 'Programa de Promoción y Prevención de la Salud en el Trabajo',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 150.0,
            'observaciones': ''
        })

    servicios.append({
        'descripcion': 'Matriz de identificación de peligros y evaluación de riesgos',
        'tipo': 'Setup',
        'cantidad': 1,
        'precio_unitario': 300.0,
        'observaciones': ''
    })

    if es_pequena_o_mas:
        servicios.append({
            'descripcion': 'Programa de prevención integral de alcohol, tabaco y otras drogas',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 200.0,
            'observaciones': ''
        })
        servicios.append({
            'descripcion': 'Programa de prevención de riesgos psicosociales',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 250.0,
            'observaciones': ''
        })
        servicios.append({
            'descripcion': 'Plan anual de capacitaciones en SST',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 180.0,
            'observaciones': ''
        })
        servicios.append({
            'descripcion': 'Plan de prevención de amenazas naturales y riesgos antrópicos',
            'tipo': 'Setup',
            'cantidad': 1,
            'precio_unitario': 220.0,
            'observaciones': ''
        })
        servicios.append({
            'descripcion': 'Programa de Vigilancia de la Salud (exámenes médicos ocupacionales)',
            'tipo': 'Especial',
            'cantidad': 1,
            'precio_unitario': 500.0,
            'observaciones': ''
        })
    else:
        servicios.append({
            'descripcion': 'Vigilancia de la salud (exámenes médicos ocupacionales)',
            'tipo': 'Especial',
            'cantidad': 1,
            'precio_unitario': 300.0,
            'observaciones': ''
        })

    # ---- MENSUAL ----
    costo_base = horas * 30 * 1.4
    servicios.append({
        'descripcion': f'Gestión de {responsable} de SST ({horas} horas/mes) - seguimiento de obligaciones',
        'tipo': 'Mensual',
        'cantidad': 1,
        'precio_unitario': round(costo_base, 2),
        'observaciones': ''
    })

    if cliente.get('tiene_grupos_prioritarios', False):
        servicios.append({
            'descripcion': 'Atención especial a grupos de atención prioritaria',
            'tipo': 'Mensual',
            'cantidad': 1,
            'precio_unitario': 60.0,
            'observaciones': ''
        })

    # ---- ESPECIAL (anual) ----
    servicios.append({
        'descripcion': 'Mediciones de higiene industrial (ruido, iluminación, agentes químicos, etc.) - anual',
        'tipo': 'Especial',
        'cantidad': 1,
        'precio_unitario': 600.0,
        'observaciones': ''
    })

    return servicios

# ------------------- ENDPOINTS API -------------------
@app.route('/')
def index():
    return render_template('acceso_cliente.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/entregables')
def entregables():
    return render_template('entregables.html')

@app.route('/clientes')
def clientes():
    return render_template('clientes.html')

@app.route('/propuestas')
def propuestas():
    return render_template('propuestas.html')

@app.route('/editar-propuesta/<int:propuesta_id>')
def editar_propuesta(propuesta_id):
    return render_template('editar_propuesta.html', propuesta_id=propuesta_id)

@app.route('/cliente/<codigo>')
def portal_cliente(codigo):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id, razon_social FROM clientes WHERE codigo_acceso=?", (codigo,))
    cliente = c.fetchone()
    conn.close()
    if not cliente:
        # Pasar también es_cliente=True para que el navbar sea el del cliente
        return render_template('cliente_error.html', mensaje='Código de acceso inválido', es_cliente=True)
    return render_template(
        'cliente_portal.html',
        cliente_id=cliente[0],
        cliente_nombre=cliente[1],
        codigo=codigo,
        es_cliente=True,
        codigo_cliente=codigo
    )
@app.route('/api/dashboard', methods=['GET'])
def api_dashboard():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM clientes")
    total_clientes = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM propuestas")
    total_propuestas = c.fetchone()[0]
    c.execute("SELECT SUM(total_anual) FROM propuestas")
    ingresos = c.fetchone()[0] or 0
    c.execute("SELECT estado, COUNT(*) FROM propuestas GROUP BY estado")
    estados = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return jsonify({
        'total_clientes': total_clientes,
        'total_propuestas': total_propuestas,
        'ingresos_totales': round(ingresos, 2),
        'estados': estados
    })

@app.route('/api/actividades', methods=['GET'])
def listar_actividades():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT codigo, descripcion, nivel_riesgo FROM actividades ORDER BY descripcion")
    rows = c.fetchall()
    conn.close()
    return jsonify([{'codigo': r[0], 'descripcion': r[1], 'nivel_riesgo': r[2]} for r in rows])

@app.route('/api/clientes', methods=['GET'])
def listar_clientes():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id, razon_social, ruc, representante, email, telefono, sector, numero_trabajadores, actividad_codigo, tiene_grupos_prioritarios, codigo_acceso FROM clientes ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return jsonify([{
        'id': r[0], 'razon_social': r[1], 'ruc': r[2], 'representante': r[3],
        'email': r[4], 'telefono': r[5], 'sector': r[6], 'numero_trabajadores': r[7],
        'actividad_codigo': r[8], 'tiene_grupos_prioritarios': bool(r[9]),
        'codigo_acceso': r[10]
    } for r in rows])

@app.route('/api/clientes', methods=['POST'])
def crear_cliente():
    data = request.json
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        codigo = generar_codigo_acceso()
        c.execute('''INSERT INTO clientes 
            (razon_social, ruc, representante, email, telefono, sector, actividad_codigo, 
             numero_trabajadores, tiene_grupos_prioritarios, fecha_registro, codigo_acceso)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (data['razon_social'], data['ruc'], data['representante'], data['email'],
             data['telefono'], data['sector'], data['actividad_codigo'],
             data['numero_trabajadores'], data.get('tiene_grupos_prioritarios', 0),
             datetime.now().isoformat(), codigo))
        conn.commit()
        cliente_id = c.lastrowid
        conn.close()
        return jsonify({'id': cliente_id, 'mensaje': 'Cliente creado', 'codigo_acceso': codigo}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/clientes/<int:cliente_id>', methods=['DELETE'])
def eliminar_cliente(cliente_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        c.execute("SELECT id FROM propuestas WHERE cliente_id=?", (cliente_id,))
        propuestas = c.fetchall()
        for p in propuestas:
            c.execute("DELETE FROM servicios WHERE propuesta_id=?", (p[0],))
        c.execute("DELETE FROM propuestas WHERE cliente_id=?", (cliente_id,))
        c.execute("DELETE FROM proyectos WHERE cliente_id=?", (cliente_id,))
        c.execute("DELETE FROM clientes WHERE id=?", (cliente_id,))
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Cliente eliminado correctamente'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/clientes/<int:cliente_id>/codigo', methods=['GET'])
def obtener_codigo_cliente(cliente_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT codigo_acceso FROM clientes WHERE id=?", (cliente_id,))
    row = c.fetchone()
    conn.close()
    if row and row[0]:
        return jsonify({'codigo': row[0]})
    return jsonify({'codigo': None}), 404

@app.route('/api/clientes/<int:cliente_id>/codigo', methods=['POST'])
def regenerar_codigo_cliente(cliente_id):
    nuevo_codigo = generar_codigo_acceso()
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("UPDATE clientes SET codigo_acceso=? WHERE id=?", (nuevo_codigo, cliente_id))
    conn.commit()
    conn.close()
    return jsonify({'codigo': nuevo_codigo})

@app.route('/api/propuestas/generar', methods=['POST'])
def generar_propuesta():
    data = request.json
    cliente_id = data.get('cliente_id')
    if not cliente_id:
        return jsonify({'error': 'Se requiere cliente_id'}), 400

    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Cliente no encontrado'}), 404

    cliente = {
        'id': row[0], 'razon_social': row[1], 'ruc': row[2], 'representante': row[3],
        'email': row[4], 'telefono': row[5], 'sector': row[6], 'actividad_codigo': row[7],
        'numero_trabajadores': row[8], 'tiene_grupos_prioritarios': row[9]
    }

    servicios = generar_servicios_base(cliente)

    total_mensual = sum(s['precio_unitario'] * s['cantidad'] for s in servicios if s['tipo'] == 'Mensual')
    total_setup = sum(s['precio_unitario'] * s['cantidad'] for s in servicios if s['tipo'] == 'Setup')
    total_especial = sum(s['precio_unitario'] * s['cantidad'] for s in servicios if s['tipo'] == 'Especial')
    total_anual = total_setup + total_mensual + total_especial

    fecha = datetime.now().isoformat()
    c.execute('''INSERT INTO propuestas 
        (cliente_id, fecha_creacion, fecha_actualizacion, estado, total_mensual, total_anual, observaciones)
        VALUES (?,?,?,?,?,?,?)''',
              (cliente_id, fecha, fecha, 'Borrador', total_mensual, total_anual, 'Generada automáticamente'))
    propuesta_id = c.lastrowid

    for s in servicios:
        subtotal = s['cantidad'] * s['precio_unitario']
        c.execute('''INSERT INTO servicios 
            (propuesta_id, descripcion, tipo, cantidad, precio_unitario, subtotal, observaciones)
            VALUES (?,?,?,?,?,?,?)''',
                  (propuesta_id, s['descripcion'], s['tipo'], s['cantidad'], s['precio_unitario'], subtotal, s.get('observaciones', '')))

    conn.commit()
    conn.close()

    return jsonify({
        'propuesta_id': propuesta_id,
        'total_mensual': total_mensual,
        'total_anual': total_anual,
        'servicios': servicios
    })

@app.route('/api/propuestas/<int:propuesta_id>', methods=['GET'])
def obtener_propuesta(propuesta_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''SELECT p.id, p.cliente_id, p.fecha_creacion, p.estado, p.total_mensual, p.total_anual, p.observaciones, p.observaciones_personalizadas,
                       c.razon_social, c.ruc, c.representante, c.email, c.telefono, c.actividad_codigo, c.numero_trabajadores
                 FROM propuestas p
                 JOIN clientes c ON p.cliente_id = c.id
                 WHERE p.id = ?''', (propuesta_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Propuesta no encontrada'}), 404

    c.execute("SELECT id, descripcion, tipo, cantidad, precio_unitario, subtotal, observaciones FROM servicios WHERE propuesta_id=?", (propuesta_id,))
    servicios = [{'id': r[0], 'descripcion': r[1], 'tipo': r[2], 'cantidad': r[3], 'precio_unitario': r[4], 'subtotal': r[5], 'observaciones': r[6] or ''} for r in c.fetchall()]
    conn.close()

    return jsonify({
        'id': row[0],
        'cliente_id': row[1],
        'fecha_creacion': row[2],
        'estado': row[3],
        'total_mensual': row[4],
        'total_anual': row[5],
        'observaciones': row[6],
        'observaciones_personalizadas': row[7],
        'cliente': {
            'razon_social': row[8],
            'ruc': row[9],
            'representante': row[10],
            'email': row[11],
            'telefono': row[12],
            'actividad_codigo': row[13],
            'numero_trabajadores': row[14]
        },
        'servicios': servicios
    })

@app.route('/api/propuestas/<int:propuesta_id>', methods=['PUT'])
def actualizar_propuesta(propuesta_id):
    data = request.json
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    
    try:
        c.execute('''UPDATE propuestas 
                     SET estado=?, observaciones=?, observaciones_personalizadas=?, fecha_actualizacion=?
                     WHERE id=?''',
                  (data.get('estado', 'Borrador'), data.get('observaciones', ''), data.get('observaciones_personalizadas', ''), datetime.now().isoformat(), propuesta_id))
        
        c.execute("SELECT id FROM servicios WHERE propuesta_id=?", (propuesta_id,))
        servicios_antiguos = [row[0] for row in c.fetchall()]
        
        c.execute("DELETE FROM servicios WHERE propuesta_id=?", (propuesta_id,))
        
        total_mensual = 0
        nuevos_servicios = []
        for s in data.get('servicios', []):
            subtotal = s['cantidad'] * s['precio_unitario']
            c.execute('''INSERT INTO servicios 
                (propuesta_id, descripcion, tipo, cantidad, precio_unitario, subtotal, observaciones)
                VALUES (?,?,?,?,?,?,?)''',
                      (propuesta_id, s['descripcion'], s['tipo'], s['cantidad'], s['precio_unitario'], subtotal, s.get('observaciones', '')))
            nuevo_id = c.lastrowid
            nuevos_servicios.append(nuevo_id)
            if s['tipo'] == 'Mensual':
                total_mensual += subtotal
        
        # Sincronizar entregables
        if servicios_antiguos:
            if nuevos_servicios:
                placeholders = ','.join(['?'] * len(nuevos_servicios))
                c.execute(f'''DELETE FROM entregables 
                             WHERE proyecto_id IN (SELECT id FROM proyectos WHERE propuesta_id=?)
                             AND servicio_id NOT IN ({placeholders})
                             AND es_autogenerado = 1''',
                          (propuesta_id,) + tuple(nuevos_servicios))
            else:
                c.execute('''DELETE FROM entregables 
                             WHERE proyecto_id IN (SELECT id FROM proyectos WHERE propuesta_id=?)
                             AND es_autogenerado = 1''',
                          (propuesta_id,))
            
            if nuevos_servicios:
                c.execute("SELECT id FROM proyectos WHERE propuesta_id=?", (propuesta_id,))
                proyectos = c.fetchall()
                for proy in proyectos:
                    c.execute("SELECT servicio_id FROM entregables WHERE proyecto_id=? AND es_autogenerado=1", (proy[0],))
                    existentes = [row[0] for row in c.fetchall()]
                    for serv_id in nuevos_servicios:
                        if serv_id not in existentes:
                            c.execute("SELECT descripcion FROM servicios WHERE id=?", (serv_id,))
                            desc = c.fetchone()[0] or 'Servicio sin descripción'
                            fecha_limite = (datetime.now() + timedelta(days=30)).isoformat()
                            c.execute('''INSERT INTO entregables 
                                (proyecto_id, servicio_id, nombre, descripcion, estado, es_autogenerado, fecha_limite)
                                VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                      (proy[0], serv_id, desc, desc, 'Pendiente', 1, fecha_limite))
        
        c.execute("SELECT SUM(subtotal) FROM servicios WHERE propuesta_id=? AND tipo='Setup'", (propuesta_id,))
        total_setup = c.fetchone()[0] or 0
        c.execute("SELECT SUM(subtotal) FROM servicios WHERE propuesta_id=? AND tipo='Especial'", (propuesta_id,))
        total_especial = c.fetchone()[0] or 0
        total_anual = total_setup + total_mensual + total_especial
        
        c.execute("UPDATE propuestas SET total_mensual=?, total_anual=? WHERE id=?", (total_mensual, total_anual, propuesta_id))
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Propuesta actualizada correctamente', 'total_mensual': total_mensual, 'total_anual': total_anual})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/propuestas', methods=['GET'])
def listar_propuestas():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''SELECT p.id, p.cliente_id, p.fecha_creacion, p.estado, p.total_mensual, p.total_anual,
                       c.razon_social, c.ruc, p.fecha_actualizacion
                 FROM propuestas p
                 JOIN clientes c ON p.cliente_id = c.id
                 ORDER BY p.id DESC''')
    rows = c.fetchall()
    conn.close()
    return jsonify([{
        'id': r[0], 'cliente_id': r[1], 'fecha_creacion': r[2], 'estado': r[3],
        'total_mensual': r[4], 'total_anual': r[5], 'cliente': r[6], 'ruc': r[7],
        'fecha_actualizacion': r[8]
    } for r in rows])

@app.route('/api/propuestas/<int:propuesta_id>', methods=['DELETE'])
def eliminar_propuesta(propuesta_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        c.execute('''DELETE FROM entregables WHERE proyecto_id IN 
                     (SELECT id FROM proyectos WHERE propuesta_id=?) AND es_autogenerado=1''', (propuesta_id,))
        c.execute("DELETE FROM servicios WHERE propuesta_id=?", (propuesta_id,))
        c.execute("DELETE FROM propuestas WHERE id=?", (propuesta_id,))
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Propuesta eliminada correctamente'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

# ------------------- ENDPOINTS PARA SEGUIMIENTO -------------------
@app.route('/api/propuestas/<int:propuesta_id>/servicios', methods=['GET'])
def obtener_servicios_propuesta(propuesta_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id, descripcion, tipo FROM servicios WHERE propuesta_id=?", (propuesta_id,))
    servicios = [{'id': r[0], 'descripcion': r[1], 'tipo': r[2]} for r in c.fetchall()]
    conn.close()
    return jsonify(servicios)

@app.route('/api/proyectos', methods=['GET'])
def listar_proyectos():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''SELECT p.id, p.cliente_id, p.propuesta_id, p.nombre, p.fecha_inicio, p.fecha_fin, p.estado, p.consultor_responsable,
                       c.razon_social, pr.estado as propuesta_estado
                 FROM proyectos p
                 JOIN clientes c ON p.cliente_id = c.id
                 LEFT JOIN propuestas pr ON p.propuesta_id = pr.id
                 ORDER BY p.id DESC''')
    rows = c.fetchall()
    conn.close()
    return jsonify([{
        'id': r[0], 'cliente_id': r[1], 'propuesta_id': r[2], 'nombre': r[3],
        'fecha_inicio': r[4], 'fecha_fin': r[5], 'estado': r[6], 'consultor': r[7],
        'cliente': r[8], 'propuesta_estado': r[9]
    } for r in rows])

@app.route('/api/proyectos', methods=['POST'])
def crear_proyecto():
    data = request.json
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        propuesta_id = data.get('propuesta_id')
        cliente_id = data.get('cliente_id')
        
        if propuesta_id and not cliente_id:
            c.execute("SELECT cliente_id FROM propuestas WHERE id=?", (propuesta_id,))
            row = c.fetchone()
            if row:
                cliente_id = row[0]
        
        c.execute('''INSERT INTO proyectos 
            (cliente_id, propuesta_id, nombre, fecha_inicio, fecha_fin, estado, consultor_responsable)
            VALUES (?,?,?,?,?,?,?)''',
                  (cliente_id, propuesta_id, data['nombre'],
                   data['fecha_inicio'], data.get('fecha_fin'), data.get('estado', 'Activo'),
                   data.get('consultor_responsable', 'Pendiente')))
        proyecto_id = c.lastrowid
        
        if propuesta_id:
            c.execute("SELECT id, descripcion FROM servicios WHERE propuesta_id=?", (propuesta_id,))
            servicios = c.fetchall()
            for serv in servicios:
                desc = serv[1] if serv[1] else 'Servicio sin descripción'
                fecha_limite = (datetime.strptime(data['fecha_inicio'], '%Y-%m-%d') + timedelta(days=30)).isoformat()
                c.execute('''INSERT INTO entregables 
                    (proyecto_id, servicio_id, nombre, descripcion, estado, es_autogenerado, fecha_limite)
                    VALUES (?,?,?,?,?,?,?)''',
                          (proyecto_id, serv[0], desc, desc, 'Pendiente', 1, fecha_limite))
        
        conn.commit()
        conn.close()
        return jsonify({'id': proyecto_id, 'mensaje': 'Proyecto creado con entregables generados'}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/proyectos/<int:proyecto_id>', methods=['DELETE'])
def eliminar_proyecto(proyecto_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        c.execute("DELETE FROM entregables WHERE proyecto_id=?", (proyecto_id,))
        c.execute("DELETE FROM proyectos WHERE id=?", (proyecto_id,))
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Proyecto eliminado correctamente'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/entregables', methods=['GET'])
def listar_entregables():
    proyecto_id = request.args.get('proyecto_id')
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    if proyecto_id:
        c.execute('''SELECT e.id, e.proyecto_id, e.servicio_id, e.nombre, e.descripcion, e.estado,
                           e.fecha_limite, e.fecha_entrega, e.archivo, e.comentarios, e.es_autogenerado,
                           p.propuesta_id
                    FROM entregables e
                    JOIN proyectos p ON e.proyecto_id = p.id
                    WHERE e.proyecto_id = ?
                    ORDER BY e.fecha_limite''', (proyecto_id,))
    else:
        c.execute('''SELECT e.id, e.proyecto_id, e.servicio_id, e.nombre, e.descripcion, e.estado,
                           e.fecha_limite, e.fecha_entrega, e.archivo, e.comentarios, e.es_autogenerado,
                           p.propuesta_id
                    FROM entregables e
                    JOIN proyectos p ON e.proyecto_id = p.id
                    ORDER BY e.fecha_limite''')
    rows = c.fetchall()
    conn.close()
    return jsonify([{
        'id': r[0], 'proyecto_id': r[1], 'servicio_id': r[2], 'nombre': r[3],
        'descripcion': r[4], 'estado': r[5], 'fecha_limite': r[6], 'fecha_entrega': r[7],
        'archivo': r[8], 'comentarios': r[9], 'es_autogenerado': bool(r[10]),
        'propuesta_id': r[11]
    } for r in rows])

@app.route('/api/entregables', methods=['POST'])
def crear_entregable():
    data = request.json
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    try:
        c.execute('''INSERT INTO entregables 
            (proyecto_id, nombre, descripcion, estado, fecha_limite, fecha_entrega, comentarios, es_autogenerado)
            VALUES (?,?,?,?,?,?,?,?)''',
            (data['proyecto_id'], data['nombre'], data.get('descripcion', ''),
             data.get('estado', 'Pendiente'), data['fecha_limite'], data.get('fecha_entrega'), data.get('comentarios', ''), 0))
        conn.commit()
        entregable_id = c.lastrowid
        conn.close()
        return jsonify({'id': entregable_id, 'mensaje': 'Entregable creado'}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/entregables/<int:entregable_id>', methods=['PUT'])
def actualizar_entregable(entregable_id):
    # Si es una solicitud multipart (con archivo)
    if request.files:
        data = request.form
        archivo = request.files.get('archivo')
    else:
        data = request.json
        archivo = None
    
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    
    try:
        campos = []
        valores = []
        
        # Campos de texto
        for key in ['nombre', 'descripcion', 'estado', 'fecha_limite', 'fecha_entrega', 'comentarios']:
            if key in data:
                campos.append(f"{key}=?")
                valores.append(data[key])
        
        # Manejar archivo
        if archivo and archivo.filename:
            if not allowed_file(archivo.filename):
                conn.close()
                return jsonify({'error': 'Tipo de archivo no permitido'}), 400
            
            filename = secure_filename(archivo.filename)
            unique_filename = f"{entregable_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            archivo.save(filepath)
            campos.append("archivo=?")
            valores.append(unique_filename)
        
        if not campos:
            conn.close()
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        
        valores.append(entregable_id)
        c.execute(f"UPDATE entregables SET {', '.join(campos)} WHERE id=?", valores)
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Entregable actualizado'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/entregables/<int:entregable_id>', methods=['DELETE'])
def eliminar_entregable(entregable_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    # Eliminar el archivo físico si existe
    c.execute("SELECT archivo FROM entregables WHERE id=?", (entregable_id,))
    row = c.fetchone()
    if row and row[0]:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], row[0])
        if os.path.exists(filepath):
            os.remove(filepath)
    
    c.execute("DELETE FROM entregables WHERE id=?", (entregable_id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Entregable eliminado'})

# ------------------- PORTAL DEL CLIENTE - API -------------------
@app.route('/api/cliente/<codigo>/entregables', methods=['GET'])
def api_cliente_entregables(codigo):
    """API para que el cliente vea sus entregables"""
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id FROM clientes WHERE codigo_acceso=?", (codigo,))
    cliente = c.fetchone()
    if not cliente:
        conn.close()
        return jsonify({'error': 'Código inválido'}), 404
    
    cliente_id = cliente[0]
    c.execute('''
        SELECT e.id, e.nombre, e.estado, e.fecha_limite, e.fecha_entrega, e.archivo,
               p.nombre as proyecto_nombre, pr.id as propuesta_id
        FROM entregables e
        JOIN proyectos p ON e.proyecto_id = p.id
        LEFT JOIN propuestas pr ON p.propuesta_id = pr.id
        WHERE p.cliente_id = ?
        ORDER BY e.fecha_limite, e.nombre
    ''', (cliente_id,))
    rows = c.fetchall()
    conn.close()
    
    entregables = []
    for row in rows:
        entregables.append({
            'id': row[0],
            'nombre': row[1],
            'estado': row[2],
            'fecha_limite': row[3],
            'fecha_entrega': row[4],
            'archivo': row[5],
            'proyecto': row[6],
            'propuesta_id': row[7],
            'tiene_archivo': bool(row[5])
        })
    return jsonify(entregables)

@app.route('/api/cliente/<codigo>/entregable/<int:entregable_id>/download', methods=['GET'])
def cliente_descargar_archivo(codigo, entregable_id):
    """Permite al cliente descargar un archivo de su entregable"""
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''
        SELECT e.archivo FROM entregables e
        JOIN proyectos p ON e.proyecto_id = p.id
        JOIN clientes cl ON p.cliente_id = cl.id
        WHERE e.id = ? AND cl.codigo_acceso = ?
    ''', (entregable_id, codigo))
    row = c.fetchone()
    conn.close()
    
    if not row or not row[0]:
        return jsonify({'error': 'Archivo no encontrado o acceso denegado'}), 404
    
    archivo_path = os.path.join(app.config['UPLOAD_FOLDER'], row[0])
    if not os.path.exists(archivo_path):
        return jsonify({'error': 'El archivo ya no está disponible'}), 404
    
    return send_file(archivo_path, as_attachment=True, download_name=row[0])

# ------------------- GENERAR PDF DE SEGUIMIENTO -------------------
@app.route('/api/entregables/pdf', methods=['GET'])
def entregables_pdf():
    proyecto_id = request.args.get('proyecto_id')
    propuesta_id = request.args.get('propuesta_id')
    cliente_id = request.args.get('cliente_id')
    
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    
    query = '''
        SELECT e.id, e.nombre, e.estado, e.fecha_limite, e.fecha_entrega, 
               e.descripcion, c.razon_social as cliente, pr.id as propuesta_id,
               p.nombre as proyecto_nombre
        FROM entregables e
        JOIN proyectos p ON e.proyecto_id = p.id
        JOIN clientes c ON p.cliente_id = c.id
        LEFT JOIN propuestas pr ON p.propuesta_id = pr.id
        WHERE 1=1
    '''
    params = []
    if proyecto_id:
        query += " AND p.id = ?"
        params.append(proyecto_id)
    if propuesta_id:
        query += " AND pr.id = ?"
        params.append(propuesta_id)
    if cliente_id:
        query += " AND c.id = ?"
        params.append(cliente_id)
    
    query += " ORDER BY e.fecha_limite, e.nombre"
    
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    
    if not rows:
        return jsonify({'error': 'No hay entregables para generar el PDF'}), 404
    
    cliente_nombre = rows[0][6] if rows[0][6] else 'N/A'
    propuesta_display = f"#{rows[0][7]}" if rows[0][7] else 'Sin propuesta asociada'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(name='Titulo', fontSize=16, leading=20, alignment=TA_CENTER, spaceAfter=20)
    estilo_cabecera = ParagraphStyle(name='Cabecera', fontSize=11, leading=14, spaceAfter=4)
    estilo_normal = styles['Normal']
    estilo_observacion = ParagraphStyle(name='Observacion', fontSize=9, leading=11, fontName='Helvetica-Oblique')
    
    elementos = []
    
    logo_path = None
    try:
        static_folder = os.path.join(os.path.dirname(__file__), 'static')
        for file in os.listdir(static_folder):
            if file.lower().endswith('.png'):
                logo_path = os.path.join(static_folder, file)
                break
    except:
        pass
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=2*inch, height=1*inch)
            img.hAlign = 'CENTER'
            elementos.append(img)
            elementos.append(Spacer(1, 0.2*inch))
        except:
            pass
    
    elementos.append(Paragraph("REPORTE DE SEGUIMIENTO DE ENTREGABLES", estilo_titulo))
    elementos.append(Paragraph(f"Cliente: {cliente_nombre}", estilo_cabecera))
    elementos.append(Paragraph(f"Propuesta: {propuesta_display}", estilo_cabecera))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", estilo_cabecera))
    elementos.append(Spacer(1, 0.3*inch))
    
    data = [["Entregable", "Estado", "Fecha Límite", "Fecha Entrega"]]
    for row in rows:
        estado = row[2] if row[2] else 'Pendiente'
        fecha_limite = row[3][:10] if row[3] else '-'
        fecha_entrega = row[4][:10] if row[4] else '-'
        data.append([
            Paragraph(row[1] if row[1] else 'Sin nombre', estilo_normal),
            Paragraph(estado, estilo_normal),
            Paragraph(fecha_limite, estilo_normal),
            Paragraph(fecha_entrega, estilo_normal)
        ])
    
    tabla = Table(data, colWidths=[3.2*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 0.3*inch))
    
    total_pendientes = sum(1 for r in rows if r[2] == 'Pendiente' or r[2] is None)
    total_progreso = sum(1 for r in rows if r[2] == 'En Progreso')
    total_entregado = sum(1 for r in rows if r[2] == 'Entregado Cliente')
    total_aprobado = sum(1 for r in rows if r[2] == 'Aprobado Cliente')
    total_registrado = sum(1 for r in rows if r[2] == 'Registrado SUT')
    total_cerrado = sum(1 for r in rows if r[2] == 'Cerrado')
    
    elementos.append(Paragraph("<b>Resumen de estados:</b>", estilo_cabecera))
    elementos.append(Paragraph(
        f"Pendientes: {total_pendientes} | En Progreso: {total_progreso} | "
        f"Entregado Cliente: {total_entregado} | Aprobado Cliente: {total_aprobado} | "
        f"Registrado SUT: {total_registrado} | Cerrado: {total_cerrado}",
        estilo_normal
    ))
    elementos.append(Spacer(1, 0.3*inch))
    elementos.append(Paragraph("Este reporte ha sido generado automáticamente desde el sistema de seguimiento de la consultoría.", estilo_observacion))
# ... (código anterior de la función descargar_pdf)

# --- SECCIÓN PARA EL CLIENTE ---
# Obtener el código de acceso del cliente
cliente_id = row[1]  # Asumiendo que row[1] es el cliente_id
conn_cliente = sqlite3.connect('sst.db')
c_cliente = conn_cliente.cursor()
c_cliente.execute("SELECT codigo_acceso FROM clientes WHERE id=?", (cliente_id,))
codigo_row = c_cliente.fetchone()
codigo_acceso = codigo_row[0] if codigo_row else 'No disponible'
conn_cliente.close()

# Agregar la información al PDF
elementos.append(Spacer(1, 0.5*inch))
elementos.append(Paragraph("<b>INFORMACIÓN PARA EL CLIENTE</b>", styles['Heading4']))
elementos.append(Paragraph(f"<b>Código de acceso:</b> {codigo_acceso}", estilo_cabecera))
elementos.append(Paragraph(f"<b>Enlace a tu portal de seguimiento:</b>", estilo_cabecera))
enlace = f"https://sst-gestor.onrender.com/cliente/{codigo_acceso}"
elementos.append(Paragraph(f"<a href='{enlace}'>{enlace}</a>", estilo_normal))
elementos.append(Paragraph("Usa este código para consultar el estado de tu proyecto y descargar documentos.", estilo_normal))

# ... (resto del código para construir el PDF)    

    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()
    
    return send_file(BytesIO(pdf), as_attachment=True, download_name=f'reporte_seguimiento_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf')

# ------------------- FUNCIÓN DESCARGA PDF PROPUESTA -------------------
@app.route('/api/propuestas/<int:propuesta_id>/pdf', methods=['GET'])
def descargar_pdf(propuesta_id):
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''SELECT p.id, p.cliente_id, p.total_mensual, p.total_anual, p.observaciones_personalizadas,
                       c.razon_social, c.ruc, c.representante, c.email, c.telefono,
                       c.actividad_codigo, c.numero_trabajadores
                 FROM propuestas p
                 JOIN clientes c ON p.cliente_id = c.id
                 WHERE p.id = ?''', (propuesta_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Propuesta no encontrada'}), 404

    c.execute("SELECT descripcion, tipo, cantidad, precio_unitario, subtotal, observaciones FROM servicios WHERE propuesta_id=?", (propuesta_id,))
    servicios = c.fetchall()
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(name='Titulo', fontSize=18, leading=22, alignment=TA_CENTER, spaceAfter=20)
    estilo_subtitulo = ParagraphStyle(name='Subtitulo', fontSize=12, leading=16, alignment=TA_CENTER, spaceAfter=15, textColor=colors.grey)
    estilo_cabecera = ParagraphStyle(name='Cabecera', fontSize=11, leading=14, spaceAfter=4)
    estilo_normal = styles['Normal']
    estilo_observacion = ParagraphStyle(name='Observacion', fontSize=9, leading=11, spaceAfter=6, fontName='Helvetica-Oblique')

    elementos = []

    logo_path = None
    try:
        static_folder = os.path.join(os.path.dirname(__file__), 'static')
        for file in os.listdir(static_folder):
            if file.lower().endswith('.png'):
                logo_path = os.path.join(static_folder, file)
                break
    except Exception:
        pass

    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=2*inch, height=1*inch)
            img.hAlign = 'CENTER'
            elementos.append(img)
            elementos.append(Spacer(1, 0.2*inch))
        except Exception:
            pass

    elementos.append(Paragraph("PROPUESTA TÉCNICA Y ECONÓMICA", estilo_titulo))
    elementos.append(Paragraph("Seguridad y Salud en el Trabajo (SST)", estilo_subtitulo))
    elementos.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y')}", estilo_cabecera))
    elementos.append(Spacer(1, 0.2*inch))

    elementos.append(Paragraph("<b>DATOS DEL CLIENTE</b>", styles['Heading4']))
    elementos.append(Paragraph(f"<b>Razón Social:</b> {row[5]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>RUC:</b> {row[6]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>Representante:</b> {row[7]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>Email:</b> {row[8]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>Teléfono:</b> {row[9]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>Actividad:</b> {row[10]}", estilo_cabecera))
    elementos.append(Paragraph(f"<b>N° Trabajadores:</b> {row[11]}", estilo_cabecera))
    elementos.append(Spacer(1, 0.3*inch))

    if row[4]:
        elementos.append(Paragraph("<b>OBSERVACIONES ESPECÍFICAS</b>", styles['Heading4']))
        elementos.append(Paragraph(str(row[4]), estilo_observacion))
        elementos.append(Spacer(1, 0.2*inch))

    elementos.append(Paragraph("<b>SERVICIOS INCLUIDOS</b>", styles['Heading4']))
    data = [["Descripción", "Tipo", "Cant.", "Precio", "Observaciones"]]

    total_setup = 0.0
    total_mensual = 0.0
    total_especial = 0.0

    for s in servicios:
        if s[1] == 'Setup':
            total_setup += s[4]
        elif s[1] == 'Mensual':
            total_mensual += s[4]
        elif s[1] == 'Especial':
            total_especial += s[4]

        desc = s[0]
        obs = s[5] if s[5] else ''

        data.append([
            Paragraph(desc, estilo_normal),
            s[1],
            str(s[2]),
            f"${s[3]:.2f}",
            Paragraph(obs, estilo_observacion)
        ])

    data.append(["SUBTOTAL SETUP:", "", "", f"${total_setup:.2f}", ""])
    data.append(["SUBTOTAL MENSUAL:", "", "", f"${total_mensual:.2f}", ""])
    data.append(["SUBTOTAL ESPECIAL:", "", "", f"${total_especial:.2f}", ""])

    total_simple = total_setup + total_mensual + total_especial
    data.append(["TOTAL:", "", "", f"${total_simple:.2f}", ""])

    tabla = Table(data, colWidths=[2.8*inch, 0.7*inch, 0.5*inch, 1.0*inch, 1.5*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0, -4), (-1, -2), colors.lightgreen),
        ('FONTNAME', (0, -4), (-1, -2), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 13),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.darkblue),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -5), 'LEFT'),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 0.3*inch))

    elementos.append(Paragraph("<b>Nota:</b>", styles['Heading4']))
    elementos.append(Paragraph("El valor TOTAL corresponde a la suma de todos los servicios. Puede ser dividido en 12 cuotas mensuales.", estilo_normal))
    elementos.append(Spacer(1, 0.1*inch))
    elementos.append(Paragraph("Esta propuesta se basa en el Decreto Ejecutivo 255 (Reglamento de Seguridad y Salud en el Trabajo) y el Acuerdo Ministerial 196 (Normas para el cumplimiento y control de obligaciones laborales en SST).", estilo_normal))
    elementos.append(Spacer(1, 0.1*inch))
    elementos.append(Paragraph("Los precios NO incluyen IVA y están sujetos a cambios según la normativa vigente.", estilo_normal))
    elementos.append(Spacer(1, 0.5*inch))

    elementos.append(Paragraph("_______________________________", styles['Normal']))
    elementos.append(Paragraph("ISOAUDITA", styles['Normal']))
    elementos.append(Paragraph("+593 994304967", styles['Normal']))
    elementos.append(Paragraph("comercial@isoaudita.com", styles['Normal']))
    elementos.append(Paragraph("QUITO - ECUADOR", styles['Normal']))

    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()

    return send_file(BytesIO(pdf), as_attachment=True, download_name=f'propuesta_{propuesta_id}.pdf')

# ------------------- EXPORTAR A EXCEL -------------------
@app.route('/api/exportar/clientes', methods=['GET'])
def exportar_clientes_excel():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute("SELECT id, razon_social, ruc, representante, email, telefono, sector, actividad_codigo, numero_trabajadores, fecha_registro, codigo_acceso FROM clientes ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ['ID', 'Razón Social', 'RUC', 'Representante', 'Email', 'Teléfono', 'Sector', 'Actividad', 'N° Trabajadores', 'Fecha Registro', 'Código Acceso']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row[0])
        ws.cell(row=r_idx, column=2, value=row[1])
        ws.cell(row=r_idx, column=3, value=row[2])
        ws.cell(row=r_idx, column=4, value=row[3])
        ws.cell(row=r_idx, column=5, value=row[4])
        ws.cell(row=r_idx, column=6, value=row[5])
        ws.cell(row=r_idx, column=7, value=row[6])
        ws.cell(row=r_idx, column=8, value=row[7])
        ws.cell(row=r_idx, column=9, value=row[8])
        ws.cell(row=r_idx, column=10, value=row[9] if row[9] else '')
        ws.cell(row=r_idx, column=11, value=row[10] if row[10] else '')

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'clientes_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/api/exportar/propuestas', methods=['GET'])
def exportar_propuestas_excel():
    conn = sqlite3.connect('sst.db')
    c = conn.cursor()
    c.execute('''SELECT p.id, c.razon_social, c.ruc, p.fecha_creacion, p.estado, p.total_mensual, p.total_anual
                 FROM propuestas p JOIN clientes c ON p.cliente_id = c.id ORDER BY p.id DESC''')
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propuestas"

    headers = ['ID', 'Cliente', 'RUC', 'Fecha', 'Estado', 'Total Mensual', 'Total Anual']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row[0])
        ws.cell(row=r_idx, column=2, value=row[1])
        ws.cell(row=r_idx, column=3, value=row[2])
        ws.cell(row=r_idx, column=4, value=row[3][:10] if row[3] else '')
        ws.cell(row=r_idx, column=5, value=row[4])
        ws.cell(row=r_idx, column=6, value=row[5])
        ws.cell(row=r_idx, column=7, value=row[6])

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'propuestas_{datetime.now().strftime("%Y%m%d")}.xlsx')

# ------------------- STATIC FILES -------------------
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

# ------------------- INICIAR APP -------------------
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)